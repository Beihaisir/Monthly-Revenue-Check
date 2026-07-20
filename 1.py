# -*- coding: utf-8 -*-
"""
Streamlit 支付方式销售报表自动汇总工具

功能：
1. 页面交互式上传文件：
   - 主表：支付方式销售报表（必传）
   - 美团外卖账单（可选）
   - 美团团购账单（可选）
   - 抖音团购账单（可选）
   - 淘宝闪购账单（可选）
   - 易宝账单（可选）

2. 文件名可以任意变化，只要表内结构/字段名一致即可。

3. 自动生成并下载处理后的 xlsx：
   - 当日金额合计
   - 各类支付累和
   - 当日易宝支付统计
   - 当日易宝到账(含手续费)
   - 当日美团外卖到账
   - 当日美团团购到账
   - 当日抖音到账
   - 当日淘宝闪购到账
   - 当日易宝支付到账差异
   - 当日美团外卖到账差异
   - 当日美团团购到账差异
   - 当日抖音团购到账差异
   - 当日淘宝闪购到账差异

运行：
    pip install streamlit pandas openpyxl xlrd
    streamlit run payment_summary_streamlit_app_v4.py

说明：
- 输入支持 .csv/.xls/.xlsx；输出统一为 .xlsx。
- 由于 Streamlit Web 环境不依赖 Microsoft Excel，本脚本不使用 pywin32。
"""

from __future__ import annotations

import io
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =========================
# 常量配置
# =========================

REQUIRED_MAIN_HEADERS = ["日期", "支付方式名称", "金额"]

COL_DAILY_TOTAL = "当日金额合计"
COL_CUMULATIVE = "各类支付累和"
COL_YIBAO = "当日易宝支付统计"
COL_YIBAO_RECEIVED = "当日易宝到账(含手续费)"
COL_MT_WAIMAI = "当日美团外卖到账"
COL_MT_TUANGOU = "当日美团团购到账"
COL_DOUYIN = "当日抖音到账"
COL_TAOBAO = "当日淘宝闪购到账"
COL_DIFF_YIBAO = "当日易宝支付到账差异"
COL_DIFF_MT_WAIMAI = "当日美团外卖到账差异"
COL_DIFF_MT_TUANGOU = "当日美团团购到账差异"
COL_DIFF_DOUYIN = "当日抖音团购到账差异"
COL_DIFF_TAOBAO = "当日淘宝闪购到账差异"

RESULT_COLUMNS = [
    COL_DAILY_TOTAL,
    COL_CUMULATIVE,
    COL_YIBAO,
    COL_YIBAO_RECEIVED,
    COL_MT_WAIMAI,
    COL_MT_TUANGOU,
    COL_DOUYIN,
    COL_TAOBAO,
    COL_DIFF_YIBAO,
    COL_DIFF_MT_WAIMAI,
    COL_DIFF_MT_TUANGOU,
    COL_DIFF_DOUYIN,
    COL_DIFF_TAOBAO,
]

DAILY_MERGE_COLUMNS = [
    COL_DAILY_TOTAL,
    COL_YIBAO,
    COL_YIBAO_RECEIVED,
    COL_MT_WAIMAI,
    COL_MT_TUANGOU,
    COL_DOUYIN,
    COL_TAOBAO,
    COL_DIFF_YIBAO,
    COL_DIFF_MT_WAIMAI,
    COL_DIFF_MT_TUANGOU,
    COL_DIFF_DOUYIN,
    COL_DIFF_TAOBAO,
]

YIBAO_PAYMENT_NAMES = {"微信支付", "微信支付(扫)", "支付宝(扫)", "支付宝支付"}

DIFF_PAYMENT_NAMES = {
    COL_DIFF_MT_WAIMAI: "美团外卖",
    COL_DIFF_MT_TUANGOU: "美团团购",
    COL_DIFF_DOUYIN: "抖音团购",
    COL_DIFF_TAOBAO: "淘宝闪购",
}

BILL_CONFIGS = {
    "meituan_waimai": {
        "display_name": "美团外卖",
        "date_header": "账单日期",
        "amount_header": "账单金额",
    },
    "meituan_tuangou": {
        "display_name": "美团团购",
        "date_header": "日期",
        "amount_header": "商家应得",
    },
    "douyin": {
        "display_name": "抖音团购",
        "date_header": "核销日期",        # ← 修正：原来是"结束日期"
        "amount_header": "商家应得",      # ← 修正：原来是"已结算金额"
    },
    "taobao": {
        "display_name": "淘宝闪购",
        "date_header": "账单日期",
        "amount_header": "结算金额",
    },
}


# =========================
# 通用解析函数
# =========================


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).replace("\u3000", " ").strip()


def normalize_header(value: Any) -> str:
    """表头匹配：去除空白、换行，降低导表格式差异影响。"""
    return normalize_text(value).replace(" ", "").replace("\n", "").replace("\r", "")


def normalize_payment_name(value: Any) -> str:
    """支付方式名称匹配：统一全角括号。"""
    return normalize_text(value).replace("（", "(").replace("）", ")")


def is_blank(value: Any) -> bool:
    return normalize_text(value) == ""


def parse_excel_date(value: Any) -> date:
    """解析 Excel 日期、序列号、文本日期。"""
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)) and not isinstance(value, bool) and not pd.isna(value):
        # Excel 日期序列，兼容 xls/xlsx 常见日期。
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()

    text = normalize_text(value)
    if not text:
        raise ValueError("日期为空")

    match = re.search(r"(20\d{2}|19\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if match:
        y, m, d = map(int, match.groups())
        return date(y, m, d)

    match = re.search(r"(20\d{2}|19\d{2})年(\d{1,2})月(\d{1,2})日", text)
    if match:
        y, m, d = map(int, match.groups())
        return date(y, m, d)

    for fmt in (
        "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日",
        "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y.%m.%d %H:%M:%S",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    raise ValueError(f"无法识别日期：{value!r}")


def parse_amount(value: Any) -> Decimal:
    """解析金额，空值按 0，负数正常保留。"""
    if value is None:
        return Decimal("0")
    if isinstance(value, float) and pd.isna(value):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return Decimal(str(value))

    text = normalize_text(value)
    if text in {"", "-", "—", "--"}:
        return Decimal("0")

    text = text.replace(",", "").replace("￥", "").replace("¥", "").replace("元", "")
    text = text.replace(" ", "")

    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"无法识别金额：{value!r}") from exc


def to_number(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def clean_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


# =========================
# Excel 读取与识别
# =========================


def _get_uploaded_name(uploaded_file) -> str:
    return getattr(uploaded_file, "name", "") or ""


def read_table_all_sheets(uploaded_file) -> Dict[str, pd.DataFrame]:
    """读取上传的 csv/xls/xlsx。返回 {sheet_name: DataFrame}，表头先不指定。"""
    uploaded_file.seek(0)
    name = _get_uploaded_name(uploaded_file).lower()

    if name.endswith(".csv"):
        last_error = None
        for enc in ("utf-8-sig", "gb18030", "gbk", "utf-8"):
            try:
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file, header=None, dtype=object, encoding=enc)
                return {"CSV": df}
            except Exception as exc:
                last_error = exc
        raise ValueError(f"CSV 文件读取失败，请检查编码或文件内容：{last_error}")

    return pd.read_excel(uploaded_file, sheet_name=None, header=None, dtype=object)


def find_header_row_and_columns(
    df: pd.DataFrame,
    required_headers: Sequence[str],
    max_scan_rows: int = 100,
    max_scan_cols: int = 200,
) -> Tuple[int, Dict[str, int]]:
    required_norm = {normalize_header(h): h for h in required_headers}
    row_limit = min(len(df), max_scan_rows)
    col_limit = min(len(df.columns), max_scan_cols)

    for row_idx in range(row_limit):
        found: Dict[str, int] = {}
        for col_idx in range(col_limit):
            text = normalize_header(df.iat[row_idx, col_idx])
            if text in required_norm:
                found[required_norm[text]] = col_idx
        if all(h in found for h in required_headers):
            return row_idx, found

    raise ValueError(f"未找到表头行，需要包含：{', '.join(required_headers)}")


def choose_main_sheet(sheets: Dict[str, pd.DataFrame]) -> Tuple[str, pd.DataFrame, int, Dict[str, int]]:
    for sheet_name, df in sheets.items():
        try:
            header_row, header_cols = find_header_row_and_columns(df, REQUIRED_MAIN_HEADERS)
            return sheet_name, df, header_row, header_cols
        except ValueError:
            continue
    raise ValueError("主表中未找到包含“日期 / 支付方式名称 / 金额”的工作表。")


def choose_bill_sheet(
    sheets: Dict[str, pd.DataFrame],
    display_name: str,
    date_header: str,
    amount_header: str,
    preferred_sheet: Optional[str] = None,
) -> Tuple[str, pd.DataFrame, int, Dict[str, int]]:
    required = [date_header, amount_header]
    
    # 如果用户指定了sheet名称，优先使用
    if preferred_sheet is not None and preferred_sheet in sheets:
        try:
            df = sheets[preferred_sheet]
            header_row, header_cols = find_header_row_and_columns(df, required)
            return preferred_sheet, df, header_row, header_cols
        except ValueError:
            pass  # 指定的sheet不匹配，继续查找其他sheet
    
    # 自动查找第一个匹配的sheet
    for sheet_name, df in sheets.items():
        try:
            header_row, header_cols = find_header_row_and_columns(df, required)
            return sheet_name, df, header_row, header_cols
        except ValueError:
            continue
    raise ValueError(f"{display_name}账单中未找到包含“{date_header} / {amount_header}”的工作表。")


def find_matching_sheets(uploaded_file, config: Dict[str, str]) -> List[Tuple[str, int, Dict[str, int]]]:
    """检测上传文件中有哪些工作表包含所需的表头字段。
    
    返回: [(sheet_name, header_row, header_cols), ...]
    """
    if uploaded_file is None:
        return []
    try:
        sheets = read_table_all_sheets(uploaded_file)
    except Exception:
        return []
    
    required = [config["date_header"], config["amount_header"]]
    matches = []
    for sheet_name, df in sheets.items():
        try:
            header_row, header_cols = find_header_row_and_columns(df, required)
            matches.append((sheet_name, header_row, header_cols))
        except ValueError:
            continue
    return matches


def remove_existing_result_columns(df: pd.DataFrame, header_row: int) -> Tuple[pd.DataFrame, List[int]]:
    targets = {normalize_header(name) for name in RESULT_COLUMNS}
    keep_cols: List[int] = []
    removed_cols: List[int] = []

    for col_idx in range(len(df.columns)):
        header_text = normalize_header(df.iat[header_row, col_idx]) if header_row < len(df) else ""
        if header_text in targets:
            removed_cols.append(col_idx)
        else:
            keep_cols.append(col_idx)

    cleaned = df.iloc[:, keep_cols].copy()
    cleaned.columns = range(len(cleaned.columns))
    return cleaned, removed_cols




def trim_trailing_blank_columns(df: pd.DataFrame, last_row: Optional[int] = None) -> pd.DataFrame:
    """删除尾部完全空白列，避免输出结果后多出空白列。"""
    if last_row is None:
        last_row = len(df) - 1
    last_col = -1
    check_df = df.iloc[: last_row + 1, :]
    for col_idx in range(len(check_df.columns)):
        if any(not is_blank(value) for value in check_df.iloc[:, col_idx].tolist()):
            last_col = col_idx
    if last_col < 0:
        return df.copy()
    trimmed = df.iloc[:, : last_col + 1].copy()
    trimmed.columns = range(len(trimmed.columns))
    return trimmed

def find_last_data_row(df: pd.DataFrame, header_row: int, key_cols: Iterable[int]) -> int:
    last = header_row
    key_cols = list(key_cols)
    for row_idx in range(header_row + 1, len(df)):
        if any(not is_blank(df.iat[row_idx, col_idx]) for col_idx in key_cols if col_idx < len(df.columns)):
            last = row_idx
    if last == header_row:
        raise ValueError("表头下方没有找到数据行。")
    return last


# =========================
# 汇总计算
# =========================


def collect_main_rows(
    df: pd.DataFrame,
    header_row: int,
    last_data_row: int,
    date_col: int,
    payment_name_col: int,
    amount_col: int,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    for row_idx in range(header_row + 1, last_data_row + 1):
        raw_date = df.iat[row_idx, date_col]
        payment_name = normalize_payment_name(df.iat[row_idx, payment_name_col])
        raw_amount = df.iat[row_idx, amount_col]

        if is_blank(raw_date) and payment_name == "" and is_blank(raw_amount):
            continue
        if is_blank(raw_date):
            raise ValueError(f"主表第 {row_idx + 1} 行日期为空，无法计算。")
        if payment_name == "":
            raise ValueError(f"主表第 {row_idx + 1} 行支付方式名称为空，无法计算。")

        rows.append({
            "row_idx": row_idx,
            "date": parse_excel_date(raw_date),
            "payment_name": payment_name,
            "amount": parse_amount(raw_amount),
        })

    if not rows:
        raise ValueError("主表没有可处理的数据行。")
    return rows


def build_main_summary_maps(rows: List[Dict[str, Any]]):
    daily_total: Dict[date, Decimal] = defaultdict(lambda: Decimal("0"))
    yibao_daily: Dict[date, Decimal] = defaultdict(lambda: Decimal("0"))
    payment_daily: Dict[str, Dict[date, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    payment_amount_by_date: Dict[Tuple[date, str], Decimal] = defaultdict(lambda: Decimal("0"))

    for item in rows:
        d = item["date"]
        p = item["payment_name"]
        a = item["amount"]
        daily_total[d] += a
        payment_daily[p][d] += a
        payment_amount_by_date[(d, p)] += a
        if p in YIBAO_PAYMENT_NAMES:
            yibao_daily[d] += a

    cumulative_by_payment_date: Dict[Tuple[str, date], Decimal] = {}
    for payment_name, date_amount_map in payment_daily.items():
        running = Decimal("0")
        for d in sorted(date_amount_map.keys()):
            running += date_amount_map[d]
            cumulative_by_payment_date[(payment_name, d)] = running

    return daily_total, cumulative_by_payment_date, yibao_daily, payment_amount_by_date


def get_date_blocks_by_current_order(rows: List[Dict[str, Any]]) -> List[Tuple[date, int, int]]:
    blocks: List[Tuple[date, int, int]] = []
    current_date = rows[0]["date"]
    start_idx = rows[0]["row_idx"]
    prev_idx = rows[0]["row_idx"]

    for item in rows[1:]:
        row_idx = item["row_idx"]
        d = item["date"]
        if d == current_date and row_idx == prev_idx + 1:
            prev_idx = row_idx
        else:
            blocks.append((current_date, start_idx, prev_idx))
            current_date = d
            start_idx = row_idx
            prev_idx = row_idx

    blocks.append((current_date, start_idx, prev_idx))
    return blocks


def read_bill_daily_map(uploaded_file, config: Dict[str, str], preferred_sheet: Optional[str] = None) -> Dict[date, Decimal]:
    sheets = read_table_all_sheets(uploaded_file)
    _, df, header_row, header_cols = choose_bill_sheet(
        sheets,
        config["display_name"],
        config["date_header"],
        config["amount_header"],
        preferred_sheet=preferred_sheet,
    )

    date_col = header_cols[config["date_header"]]
    amount_col = header_cols[config["amount_header"]]
    last_row = find_last_data_row(df, header_row, [date_col, amount_col])

    result: Dict[date, Decimal] = defaultdict(lambda: Decimal("0"))
    for row_idx in range(header_row + 1, last_row + 1):
        raw_date = df.iat[row_idx, date_col]
        raw_amount = df.iat[row_idx, amount_col]

        if is_blank(raw_date) and is_blank(raw_amount):
            continue
        if is_blank(raw_date):
            raise ValueError(f"{config['display_name']}账单第 {row_idx + 1} 行日期为空。")

        result[parse_excel_date(raw_date)] += parse_amount(raw_amount)

    return dict(result)




def read_yibao_daily_map(uploaded_file) -> Dict[date, Decimal]:
    """易宝账单：当日易宝到账(含手续费) = 收单收入(元) + 当日手续费(元) - 退款支出(元)。"""
    sheets = read_table_all_sheets(uploaded_file)
    required = ["记账日期", "业务类型", "收入(元)", "支出(元)", "手续费(元)"]

    chosen = None
    for sheet_name, df in sheets.items():
        try:
            header_row, header_cols = find_header_row_and_columns(df, required)
            chosen = (sheet_name, df, header_row, header_cols)
            break
        except ValueError:
            continue
    if chosen is None:
        raise ValueError("易宝账单中未找到包含“记账日期 / 业务类型 / 收入(元) / 支出(元) / 手续费(元)”的工作表或 CSV 表头。")

    _, df, header_row, header_cols = chosen
    date_col = header_cols["记账日期"]
    type_col = header_cols["业务类型"]
    income_col = header_cols["收入(元)"]
    expense_col = header_cols["支出(元)"]
    fee_col = header_cols["手续费(元)"]
    last_row = find_last_data_row(df, header_row, [date_col, type_col, income_col, expense_col, fee_col])

    result: Dict[date, Decimal] = defaultdict(lambda: Decimal("0"))
    for row_idx in range(header_row + 1, last_row + 1):
        raw_date = df.iat[row_idx, date_col]
        biz_type = normalize_text(df.iat[row_idx, type_col])
        if (
            is_blank(raw_date)
            and biz_type == ""
            and is_blank(df.iat[row_idx, income_col])
            and is_blank(df.iat[row_idx, expense_col])
            and is_blank(df.iat[row_idx, fee_col])
        ):
            continue
        if is_blank(raw_date):
            raise ValueError(f"易宝账单第 {row_idx + 1} 行记账日期为空。")

        d = parse_excel_date(raw_date)

        # “含手续费”按用户确认口径：
        # 当日易宝到账(含手续费) = 当日收单收入(元) + 当日手续费(元) - 当日退款支出(元)。
        # 因此手续费按记账日期逐行计入；收单行取收入，退款行取支出扣减，结算等其他类型只计手续费。
        result[d] += parse_amount(df.iat[row_idx, fee_col])
        if biz_type == "收单":
            result[d] += parse_amount(df.iat[row_idx, income_col])
        elif biz_type == "退款":
            result[d] -= parse_amount(df.iat[row_idx, expense_col])

    return dict(result)

def build_result_values(
    rows: List[Dict[str, Any]],
    bill_daily_maps: Dict[str, Dict[date, Decimal]],
) -> Tuple[Dict[int, Dict[str, Decimal]], List[Tuple[date, int, int]], Dict[str, Any]]:
    daily_total, cumulative_by_payment_date, yibao_daily, payment_amount_by_date = build_main_summary_maps(rows)
    date_blocks = get_date_blocks_by_current_order(rows)

    row_values: Dict[int, Dict[str, Decimal]] = defaultdict(dict)

    # 每行写“各类支付累和”。
    for item in rows:
        row_idx = item["row_idx"]
        key = (item["payment_name"], item["date"])
        row_values[row_idx][COL_CUMULATIVE] = cumulative_by_payment_date[key]

    # 当日类和差异类：日期区块第一行写值，后续通过 Excel 合并显示。
    for d, start_idx, end_idx in date_blocks:
        yibao_received = bill_daily_maps["yibao"].get(d, Decimal("0"))
        mt_waimai_in = bill_daily_maps["meituan_waimai"].get(d, Decimal("0"))
        mt_tuangou_in = bill_daily_maps["meituan_tuangou"].get(d, Decimal("0"))
        douyin_in = bill_daily_maps["douyin"].get(d, Decimal("0"))
        taobao_in = bill_daily_maps["taobao"].get(d, Decimal("0"))

        row_values[start_idx][COL_DAILY_TOTAL] = daily_total.get(d, Decimal("0"))
        yibao_stat = yibao_daily.get(d, Decimal("0"))
        row_values[start_idx][COL_YIBAO] = yibao_stat
        row_values[start_idx][COL_YIBAO_RECEIVED] = yibao_received
        row_values[start_idx][COL_MT_WAIMAI] = mt_waimai_in
        row_values[start_idx][COL_MT_TUANGOU] = mt_tuangou_in
        row_values[start_idx][COL_DOUYIN] = douyin_in
        row_values[start_idx][COL_TAOBAO] = taobao_in

        mt_waimai_main = payment_amount_by_date.get((d, DIFF_PAYMENT_NAMES[COL_DIFF_MT_WAIMAI]), Decimal("0"))
        mt_tuangou_main = payment_amount_by_date.get((d, DIFF_PAYMENT_NAMES[COL_DIFF_MT_TUANGOU]), Decimal("0"))
        douyin_main = payment_amount_by_date.get((d, DIFF_PAYMENT_NAMES[COL_DIFF_DOUYIN]), Decimal("0"))
        taobao_main = payment_amount_by_date.get((d, DIFF_PAYMENT_NAMES[COL_DIFF_TAOBAO]), Decimal("0"))

        row_values[start_idx][COL_DIFF_YIBAO] = yibao_stat - yibao_received
        row_values[start_idx][COL_DIFF_MT_WAIMAI] = mt_waimai_main - mt_waimai_in
        row_values[start_idx][COL_DIFF_MT_TUANGOU] = mt_tuangou_main - mt_tuangou_in
        row_values[start_idx][COL_DIFF_DOUYIN] = douyin_main - douyin_in
        row_values[start_idx][COL_DIFF_TAOBAO] = taobao_main - taobao_in

    summary = {
        "date_count": len({item["date"] for item in rows}),
        "row_count": len(rows),
        "start_date": min(item["date"] for item in rows),
        "end_date": max(item["date"] for item in rows),
    }
    return row_values, date_blocks, summary


# =========================
# 输出 xlsx
# =========================


def apply_basic_styles(ws, header_excel_row: int, last_data_excel_row: int, result_start_col: int, result_end_col: int) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    # 表头样式。
    for row in ws.iter_rows(min_row=header_excel_row, max_row=header_excel_row):
        for cell in row:
            if cell.value is not None:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                if cell.column >= result_start_col:
                    cell.fill = header_fill

    # 数据区样式。
    for row in ws.iter_rows(min_row=header_excel_row + 1, max_row=last_data_excel_row):
        for cell in row:
            if cell.value is not None:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 新增金额列格式。
    for col_idx in range(result_start_col, result_end_col + 1):
        for row_idx in range(header_excel_row + 1, last_data_excel_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    # 列宽。
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        if col_idx < result_start_col:
            ws.column_dimensions[letter].width = min(max(ws.column_dimensions[letter].width or 10, 12), 24)
        else:
            ws.column_dimensions[letter].width = 18

    # 差异列略宽。
    for col_idx in range(result_start_col, result_end_col + 1):
        header = ws.cell(row=header_excel_row, column=col_idx).value
        if header and "差异" in str(header):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

    ws.freeze_panes = ws.cell(row=header_excel_row + 1, column=1)


def create_output_workbook(
    main_df: pd.DataFrame,
    header_row: int,
    last_data_row: int,
    row_values: Dict[int, Dict[str, Decimal]],
    date_blocks: List[Tuple[date, int, int]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "处理结果"

    original_col_count = len(main_df.columns)
    result_col_map = {name: original_col_count + idx + 1 for idx, name in enumerate(RESULT_COLUMNS)}  # 1-based

    # 写原表：从第 1 行写到数据最后一行，保留标题行、导出信息行、原始表头和原始数据。
    for row_idx in range(0, last_data_row + 1):
        for col_idx in range(original_col_count):
            ws.cell(row=row_idx + 1, column=col_idx + 1, value=clean_cell_value(main_df.iat[row_idx, col_idx]))

    # 写新增表头。
    header_excel_row = header_row + 1
    for name, col_idx in result_col_map.items():
        ws.cell(row=header_excel_row, column=col_idx, value=name)

    # 写新增数据。
    for row_idx, value_map in row_values.items():
        excel_row = row_idx + 1
        for name, value in value_map.items():
            ws.cell(row=excel_row, column=result_col_map[name], value=to_number(value))

    # 合并“当日xxx”和“差异”列。各类支付累和不合并。
    for _, start_idx, end_idx in date_blocks:
        start_excel_row = start_idx + 1
        end_excel_row = end_idx + 1
        if end_excel_row <= start_excel_row:
            continue
        for name in DAILY_MERGE_COLUMNS:
            col_idx = result_col_map[name]
            ws.merge_cells(start_row=start_excel_row, start_column=col_idx, end_row=end_excel_row, end_column=col_idx)
            ws.cell(row=start_excel_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

    apply_basic_styles(
        ws,
        header_excel_row=header_excel_row,
        last_data_excel_row=last_data_row + 1,
        result_start_col=result_col_map[RESULT_COLUMNS[0]],
        result_end_col=result_col_map[RESULT_COLUMNS[-1]],
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def process_files(
    main_file,
    meituan_waimai_file,
    meituan_tuangou_file,
    douyin_file,
    taobao_file,
    yibao_file,
    sheet_selections: Optional[Dict[str, str]] = None,
) -> Tuple[bytes, Dict[str, Any]]:
    # 主表
    main_sheets = read_table_all_sheets(main_file)
    main_sheet_name, main_df, header_row, header_cols = choose_main_sheet(main_sheets)

    # 删除旧结果列，防止重复运行后多出重复字段。
    main_df, _ = remove_existing_result_columns(main_df, header_row)
    header_row, header_cols = find_header_row_and_columns(main_df, REQUIRED_MAIN_HEADERS)

    date_col = header_cols["日期"]
    payment_name_col = header_cols["支付方式名称"]
    amount_col = header_cols["金额"]
    last_data_row = find_last_data_row(main_df, header_row, [date_col, payment_name_col, amount_col])
    main_df = trim_trailing_blank_columns(main_df, last_data_row)
    header_row, header_cols = find_header_row_and_columns(main_df, REQUIRED_MAIN_HEADERS)
    date_col = header_cols["日期"]
    payment_name_col = header_cols["支付方式名称"]
    amount_col = header_cols["金额"]
    last_data_row = find_last_data_row(main_df, header_row, [date_col, payment_name_col, amount_col])
    rows = collect_main_rows(main_df, header_row, last_data_row, date_col, payment_name_col, amount_col)

    # 外部账单均为可选：未上传时使用空映射，对应到账金额按 0.00 计算。
    sheet_selections = sheet_selections or {}
    
    bill_daily_maps = {
        "meituan_waimai": (
            read_bill_daily_map(meituan_waimai_file, BILL_CONFIGS["meituan_waimai"], sheet_selections.get("meituan_waimai"))
            if meituan_waimai_file is not None else {}
        ),
        "meituan_tuangou": (
            read_bill_daily_map(meituan_tuangou_file, BILL_CONFIGS["meituan_tuangou"], sheet_selections.get("meituan_tuangou"))
            if meituan_tuangou_file is not None else {}
        ),
        "douyin": (
            read_bill_daily_map(douyin_file, BILL_CONFIGS["douyin"], sheet_selections.get("douyin"))
            if douyin_file is not None else {}
        ),
        "taobao": (
            read_bill_daily_map(taobao_file, BILL_CONFIGS["taobao"], sheet_selections.get("taobao"))
            if taobao_file is not None else {}
        ),
        "yibao": read_yibao_daily_map(yibao_file) if yibao_file is not None else {},
    }

    row_values, date_blocks, summary = build_result_values(rows, bill_daily_maps)
    output_bytes = create_output_workbook(main_df, header_row, last_data_row, row_values, date_blocks)

    summary["main_sheet_name"] = main_sheet_name
    summary["bill_date_counts"] = {key: len(value) for key, value in bill_daily_maps.items()}
    return output_bytes, summary


# =========================
# Streamlit 页面
# =========================


def render_app() -> None:
    import streamlit as st

    st.set_page_config(page_title="支付方式销售报表汇总工具", layout="wide")

    st.title("支付方式销售报表汇总工具")
    st.caption("主表必须上传；其他账单表可选。文件名可以不同；只要表内字段结构一致，即可上传处理。输入支持 .csv/.xls/.xlsx，输出为 .xlsx。")

    # ===== 各报表获取方式说明 =====
    with st.expander("📋 各报表获取方式（点击展开）", expanded=True):
        st.markdown("""
| 报表名称 | 获取路径 | 关键字段 | 数据说明 |
|---------|---------|---------|---------|
| **支付方式销售报表（主表）** | 门店管理后台 > 统计报表 > 销售日报-分支付 | 日期、支付方式名称、金额 | 门店每日各支付方式的收款汇总，作为对账主数据源 |
| **易宝账单** | 账户中心 > 资金账单 > 自助生成资金账单 > 生成汇总资金账单 | 记账日期、业务类型、收入(元)、支出(元)、手续费(元) | 易宝支付渠道的收单、退款、手续费明细 |
| **美团外卖账单** | 美团外卖\|商家中心 > 财务管理 > 下载专区 > 选择账单 | 账单日期、账单金额 | 美团外卖平台每日结算金额 |
| **美团团购账单** | 美团经营宝 > 财务管理 > 账单管理 > 每日收益 > 按日期汇总 | 日期、商家应得 | 美团团购平台每日商家应得金额 |
| **淘宝闪购账单** | 淘宝闪购商家版 > 财务管理 > 财务下载 > 账单-汇总 | 账单日期、结算金额 | 淘宝闪购平台每日结算金额 |
| **抖音团购账单** | 抖音来客后台 > 资金财务 > 财务首页 > 按日期 | 核销日期、商家应得 | 抖音团购平台每日核销后的商家应得金额 |
        """)

    with st.expander("处理规则", expanded=False):
        st.markdown(
            """
- 主表自动识别字段：`日期`、`支付方式名称`、`金额`。
- `当日金额合计`：按日期汇总所有支付方式金额，同一天合并显示一次。
- `各类支付累和`：按支付方式名称，从最早日期自然累计到当前日期，每行显示。
- `当日易宝支付统计`：`微信支付 + 微信支付(扫) + 支付宝(扫) + 支付宝支付`。
- 美团外卖：按 `账单日期` 匹配，取 `账单金额`。
- 美团团购：按 `日期` 匹配，取 `商家应得`。
- 抖音团购：按 `核销日期` 匹配，取 `商家应得`。
- 淘宝闪购：按 `账单日期` 匹配，取 `结算金额`，同一天多行全部合计，负数正常参与。
- 易宝账单：按 `记账日期` 匹配，`当日易宝到账(含手续费) = 收单收入(元) + 当日手续费(元) - 退款支出(元)`；手续费按 `记账日期` 当日合计。
- 外部账单未上传或某天无数据时，对应到账金额按 `0.00` 计算。
- `当日易宝支付到账差异 = 当日易宝支付统计 - 当日易宝到账(含手续费)`；其他差异 = 主表对应支付方式当日金额 - 外部到账金额。
            """
        )

    col1, col2 = st.columns(2)
    # ... 后续代码不变 ...
    with col1:
        main_file = st.file_uploader("1. 上传主表：支付方式销售报表（必传）", type=["csv", "xls", "xlsx"], key="main")
        meituan_waimai_file = st.file_uploader("2. 上传美团外卖账单（可选）", type=["csv", "xls", "xlsx"], key="mt_waimai")
        meituan_tuangou_file = st.file_uploader("3. 上传美团团购账单（可选）", type=["csv", "xls", "xlsx"], key="mt_tuangou")
    with col2:
        douyin_file = st.file_uploader("4. 上传抖音团购账单（可选）", type=["csv", "xls", "xlsx"], key="douyin")
        taobao_file = st.file_uploader("5. 上传淘宝闪购账单（可选）", type=["csv", "xls", "xlsx"], key="taobao")
        yibao_file = st.file_uploader("6. 上传易宝账单（可选）", type=["csv", "xls", "xlsx"], key="yibao")

    # ===== 多Sheet选择器 =====
    # 检测每个上传的文件是否有多个匹配的工作表，如果有则让用户选择
    sheet_selections: Dict[str, str] = {}
    
    # 需要检测的文件配置
    bill_files = {
        "meituan_waimai": ("美团外卖账单", meituan_waimai_file),
        "meituan_tuangou": ("美团团购账单", meituan_tuangou_file),
        "douyin": ("抖音团购账单", douyin_file),
        "taobao": ("淘宝闪购账单", taobao_file),
    }
    
    for bill_key, (display_name, uploaded_file) in bill_files.items():
        if uploaded_file is not None:
            matches = find_matching_sheets(uploaded_file, BILL_CONFIGS[bill_key])
            if len(matches) > 1:
                sheet_names = [m[0] for m in matches]
                selected = st.selectbox(
                    f"【{display_name}】检测到多个匹配的工作表，请选择：",
                    options=sheet_names,
                    key=f"sheet_select_{bill_key}",
                )
                sheet_selections[bill_key] = selected
                st.info(f"已选择「{display_name}」使用工作表：{selected}")
            elif len(matches) == 1:
                sheet_selections[bill_key] = matches[0][0]
            else:
                st.warning(f"【{display_name}】未找到包含所需字段的工作表，将按 0.00 处理。")

    if main_file is None:
        st.info("请先上传主表。其他账单表可不上传；未上传的账单，对应到账金额会按 0.00 处理。")
        return

    optional_files = {
        "美团外卖账单": meituan_waimai_file,
        "美团团购账单": meituan_tuangou_file,
        "抖音团购账单": douyin_file,
        "淘宝闪购账单": taobao_file,
        "易宝账单": yibao_file,
    }
    missing_optional = [name for name, file in optional_files.items() if file is None]
    if missing_optional:
        st.warning(
            "以下可选账单未上传，将按 0.00 计算对应到账金额：" + "、".join(missing_optional)
        )

    if st.button("生成处理后的报表", type="primary"):
        try:
            with st.spinner("正在读取、汇总并生成 Excel..."):
                output_bytes, summary = process_files(
                    main_file,
                    meituan_waimai_file,
                    meituan_tuangou_file,
                    douyin_file,
                    taobao_file,
                    yibao_file,
                    sheet_selections=sheet_selections,
                )

            st.success("处理完成，可以下载结果文件。")
            st.write(
                f"主表工作表：{summary['main_sheet_name']}；"
                f"数据行数：{summary['row_count']}；"
                f"日期范围：{summary['start_date']} 至 {summary['end_date']}；"
                f"日期天数：{summary['date_count']}。"
            )
            st.download_button(
                label="下载处理后报表.xlsx",
                data=output_bytes,
                file_name="处理后报表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            st.error("处理失败，请检查上传文件的表头字段是否和规则一致。")
            st.exception(exc)


if __name__ == "__main__":
    render_app()